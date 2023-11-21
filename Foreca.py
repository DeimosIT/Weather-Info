from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from progress.bar import IncrementalBar
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re
import pandas as pd
from openpyxl import load_workbook
from datetime import date
class City:
    def __init__(self, name, url):
        self.name = name
        self.url = url

    def getName(self):
        return self.name
    
    def getUrl(self):
        return self.url

"""def get_url(url):
    options = Options()
    #option = webdriver.FirefoxOptions()
    #options.set_preference('dom.webdriver.enabled', False)
    #options.add_argument('--headless')
    fp = webdriver.FirefoxProfile()
    fp.set_preference("network.cookie.cookieBehavior", 2)
    driver = webdriver.Firefox(options=options)
    driver.get(url)
    
    return driver"""

def get_url2(url):
    option = Options()
    option.add_argument('--headless=new')
    option.page_load_strategy = 'none'
    option.add_experimental_option("prefs", {"profile.managed_default_content_settings.images": 2})
    option.add_experimental_option('excludeSwitches', ['enable-logging'])
    driver = webdriver.Chrome(options=option)
    driver.get(url)

    return driver

def slice_list1(list):
    new_list = []
    for hour in list:
        new_list.append(hour.text)
    new_list.reverse()
    start = 0
    stop = 4
    slice_obj = slice(start, stop)
    result_list = new_list[slice_obj]
    return result_list

def slice_list2(list):
    new_list = []
    for hour in list:
        new_list.append(hour.text)
    start = 0
    stop = 9
    slice_obj = slice(start, stop)
    result_list = new_list[slice_obj]
    return result_list

def get_wind(list_wind, list_direct):

    maxWind = 0
    maxDirect = ' '
    for wind, direct in zip(list_wind, list_direct):
        if(wind > maxWind):
            maxWind = wind
            maxDirect = direct
    stringMax = str(maxWind) + str(maxDirect)
    return str.lower(stringMax)

def osadki_rain(list):
    rain_total = []
    for rain in list:
        string_vr = str(rain)
        num = re.findall(r'[-+]?(?:\d*\.*\d*.мм)', string_vr)
        string_vr2 = str(num)
        num2 = re.findall(r'[-+]?(?:\d*\.*\d+)', string_vr2)
        rain_total.append(float(num2[0]))
    result = round(sum(rain_total), 2)
    
    result_total = round(result, 3)
    return result_total

def get_num(list_str):
    new_list = []
    for num in list_str:
        string_num = str(num)
        number = re.findall(r'\d+', string_num)
        new_list.append(int(number[0]))
    return new_list


def rus_direct(list_str):
    new_string = ''
    rus_str = []
    for c in list_str:
        new_string = str(c)
        if(new_string == 'N'):
            rus_str.append('с')
        if(new_string == 'NE'):
            rus_str.append('с-в')
        if(new_string == 'E'):
            rus_str.append('в')
        if(new_string == 'SE'):
            rus_str.append('ю-в')
        if(new_string == 'S'):
            rus_str.append('ю')
        if(new_string == 'SW'):
            rus_str.append('ю-з')
        if(new_string == 'W'):
            rus_str.append('з')
        if(new_string == 'NW'):
            rus_str.append('с-з')
    return rus_str

def table_panda(listinfo, name, check):
    df = pd.DataFrame(listinfo, columns=["Min", "Max", "Нv", "Дv", "Ос.Н", "Ос.Д"], index=["Сутки 1", "Сутки 2", "Сутки 3"])
    return df

def multiple_dfs(df_list, sheets, file_name, spaces, list_name):
    with pd.ExcelWriter(file_name, mode='w') as writer:
        
        row = 0
        for dataframe, name in zip(df_list, list_name):
            row +=1
            dataframe.to_excel(writer,sheet_name=sheets,startrow=row , startcol=0)   
            row = row + len(dataframe.index) + spaces + 1
        #writer._save()

def int_name(file_name, sheets, list_name):
    wb = load_workbook(file_name)
    today = date.today()
    date_3 = today.strftime("%d.%m.%y")    
    ws = wb[sheets]
    row = 1
    coll = 1
    for name in list_name:
        ws.cell(row,coll,name)
        ws.cell(row, 7, date_3)
        ws.cell(row, 6, 'Сегодня')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 6

    wb.save(file_name)    
    #wb.close()

def get_foreca(url):
    driver = get_url2(url)
    WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="qc-cmp2-ui"]/div[2]/div/button[2]'))).click()
    if WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.XPATH, '//span[@class = "t"]'))):
        temper = driver.find_elements(By.XPATH, ('//span[@class = "t"]'))
        osadki = driver.find_elements(By.XPATH, ('//span[@class = "value rain rain_mm"]'))
        wind = driver.find_elements(By.XPATH, ('//div/span[(@class = "value wind wind_ms")]'))
        direct_wind = driver.find_elements(By.XPATH, ('//div[@class = "wind"]/img'))
        times = driver.find_elements(By.XPATH, ('//div/span[@class = "value time time_24h"]'))



    list_temp_d1 = []
    for t in temper:
        list_temp_d1.append(t.text)
    list_osadki_d1 = []
    for o in osadki:
        list_osadki_d1.append(o.text)
    list_wind_d1 = []
    for w in wind:
        list_wind_d1.append(w.text)
    list_direct_d1 = []
    for d in direct_wind:
        list_direct_d1.append(d.get_attribute('title'))
    list_time_d1 = []
    for t in times:
        list_time_d1.append(t.text)

    WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, '//div[(@class = "link next")]'))).click()

    if WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.XPATH, '//span[@class = "t"]'))):
        temper = driver.find_elements(By.XPATH, ('//span[@class = "t"]'))
        osadki = driver.find_elements(By.XPATH, ('//span[@class = "value rain rain_mm"]'))
        wind = driver.find_elements(By.XPATH, ('//div/span[(@class = "value wind wind_ms")]'))
        direct_wind = driver.find_elements(By.XPATH, ('//div[@class = "wind"]/img'))
        times = driver.find_elements(By.XPATH, ('//div/span[@class = "value time time_24h"]'))

    list_temp_d2 = []
    for t in temper:
        list_temp_d2.append(t.text)
    list_osadki_d2 = []
    for o in osadki:
        list_osadki_d2.append(o.text)
    list_wind_d2 = []
    for w in wind:
        list_wind_d2.append(w.text)
    list_direct_d2 = []
    for d in direct_wind:
        list_direct_d2.append(d.get_attribute('title'))
    list_time_d2 = []
    for t in times:
        list_time_d2.append(t.text)

    WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, '//div[(@class = "link next")]'))).click()

    if WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.XPATH, '//span[@class = "t"]'))):
        temper = driver.find_elements(By.XPATH, ('//span[@class = "t"]'))
        osadki = driver.find_elements(By.XPATH, ('//span[@class = "value rain rain_mm"]'))
        wind = driver.find_elements(By.XPATH, ('//div/span[(@class = "value wind wind_ms")]'))
        direct_wind = driver.find_elements(By.XPATH, ('//div[@class = "wind"]/img'))
        times = driver.find_elements(By.XPATH, ('//div/span[@class = "value time time_24h"]'))

    list_temp_d3 = []
    for t in temper:
        list_temp_d3.append(t.text)
    list_osadki_d3 = []
    for o in osadki:
        list_osadki_d3.append(o.text)
    list_wind_d3 = []
    for w in wind:
        list_wind_d3.append(w.text)
    list_direct_d3 = []
    for d in direct_wind:
        list_direct_d3.append(d.get_attribute('title'))
    list_time_d3 = []
    for t in times:
        list_time_d3.append(t.text)
    
    WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, '//div[(@class = "link next")]'))).click()

    if WebDriverWait(driver, 120).until(EC.visibility_of_element_located((By.XPATH, '//span[@class = "t"]'))):
        temper = driver.find_elements(By.XPATH, ('//span[@class = "t"]'))
        osadki = driver.find_elements(By.XPATH, ('//span[@class = "value rain rain_mm"]'))
        wind = driver.find_elements(By.XPATH, ('//div/span[(@class = "value wind wind_ms")]'))
        direct_wind = driver.find_elements(By.XPATH, ('//div[@class = "wind"]/img'))
        times = driver.find_elements(By.XPATH, ('//div/span[@class = "value time time_24h"]'))

    list_temp_d4 = []
    for t in temper:
        list_temp_d4.append(t.text)
    list_osadki_d4 = []
    for o in osadki:
        list_osadki_d4.append(o.text)
    list_wind_d4 = []
    for w in wind:
        list_wind_d4.append(w.text)
    list_direct_d4 = []
    for d in direct_wind:
        list_direct_d4.append(d.get_attribute('title'))
    list_time_d4 = []
    for t in times:
        list_time_d4.append(t.text)

    driver.close()
    driver.quit()

    night_temper_1 = []
    night_temper_2 = []
    night_temper_3 = []
    night_osadki_1 = []
    night_osadki_2 = []
    night_osadki_3 = []
    night_wind_1 = []
    night_wind_2 = []
    night_wind_3 = []
    night_direct_1 = []
    night_direct_2 = []
    night_direct_3 = []

    day_temper_1 = []
    day_temper_2 = []
    day_temper_3 = []
    day_osadki_1 = []
    day_osadki_2 = []
    day_osadki_3 = []
    day_wind_1 = []
    day_wind_2 = []
    day_wind_3 = []
    day_direct_1 = []
    day_direct_2 = []
    day_direct_3 = []

    for temp, o, w, d, times in zip(list_temp_d1, list_osadki_d1, list_wind_d1, list_direct_d1, list_time_d1):
        if(int(times) >= 20):
            night_temper_1.append(int(temp))
            night_osadki_1.append(o)
            night_wind_1.append(w)
            night_direct_1.append(d)
    for temp, o, w, d, times in zip(list_temp_d2, list_osadki_d2, list_wind_d2, list_direct_d2, list_time_d2):
        if(int(times) <= 8):
            night_temper_1.append(int(temp))
            night_osadki_1.append(o)
            night_wind_1.append(w)
            night_direct_1.append(d)
        if(int(times) >= 20):
            night_temper_2.append(int(temp))
            night_osadki_2.append(o)
            night_wind_2.append(w)
            night_direct_2.append(d)
        if(int(times) >= 9 and int(times) <= 19):
            day_temper_1.append(int(temp))
            day_osadki_1.append(o)
            day_wind_1.append(w)
            day_direct_1.append(d)
    for temp, o, w, d, times in zip(list_temp_d3, list_osadki_d3, list_wind_d3, list_direct_d3, list_time_d3):
        if(int(times) <= 8):
            night_temper_2.append(int(temp))
            night_osadki_2.append(o)
            night_wind_2.append(w)
            night_direct_2.append(d)
        if(int(times) >= 20):
            night_temper_3.append(int(temp))
            night_osadki_3.append(o)
            night_wind_3.append(w)
            night_direct_3.append(d)
        if(int(times) >= 9 and int(times) <= 19):
            day_temper_2.append(int(temp))
            day_osadki_2.append(o)
            day_wind_2.append(w)
            day_direct_2.append(d)
    for temp, o, w, d, times in zip(list_temp_d4, list_osadki_d4, list_wind_d4, list_direct_d4, list_time_d4):
        if(int(times) <= 8):
            night_temper_3.append(int(temp))
            night_osadki_3.append(o)
            night_wind_3.append(w)
            night_direct_3.append(d)
        if(int(times) >= 9 and int(times) <= 19):
            day_temper_3.append(int(temp))
            day_osadki_3.append(o)
            day_wind_3.append(w)
            day_direct_3.append(d)

            

    total_osadki_night_1 = osadki_rain(night_osadki_1)
    total_osadki_night_2 = osadki_rain(night_osadki_2)
    total_osadki_night_3 = osadki_rain(night_osadki_3)
    total_osadki_day_1 = osadki_rain(day_osadki_1)
    total_osadki_day_2 = osadki_rain(day_osadki_2)
    total_osadki_day_3 = osadki_rain(day_osadki_3)
    
    new_night_wind_1 = get_num(night_wind_1)
    new_night_wind_2 = get_num(night_wind_2)
    new_night_wind_3 = get_num(night_wind_3)
    new_day_wind_1 = get_num(day_wind_1)
    new_day_wind_2 = get_num(day_wind_2)
    new_day_wind_3 = get_num(day_wind_3)
    new_night_direct_1 = rus_direct(night_direct_1)
    new_night_direct_2 = rus_direct(night_direct_2)
    new_night_direct_3 = rus_direct(night_direct_3)
    new_day_direct_1 = rus_direct(day_direct_1)
    new_day_direct_2 = rus_direct(day_direct_2)
    new_day_direct_3 = rus_direct(day_direct_3)



    min_night1 = min(night_temper_1)
    min_night2 = min(night_temper_2)
    min_night3 = min(night_temper_3)
    max_day1 = max(day_temper_1)
    max_day2 = max(day_temper_2)
    max_day3 = max(day_temper_3)   
    max_wind_night_1 = get_wind(new_night_wind_1, new_night_direct_1)
    max_wind_night_2 = get_wind(new_night_wind_2, new_night_direct_2)
    max_wind_night_3 = get_wind(new_night_wind_3, new_night_direct_3)
    max_wind_day_1 = get_wind(new_day_wind_1, new_day_direct_1)
    max_wind_day_2 = get_wind(new_day_wind_2, new_day_direct_2)
    max_wind_day_3 = get_wind(new_day_wind_3, new_day_direct_3)

    Day_1 = {'Max' : max_day1,'Min' : min_night1, 'Нv' : max_wind_night_1, 'Дv': max_wind_day_1, 'Ос.Н' : total_osadki_night_1, 'Ос.Д' : total_osadki_day_1}
    Day_2 = {'Max' : max_day2,'Min' : min_night2, 'Нv' : max_wind_night_2, 'Дv': max_wind_day_2, 'Ос.Н' : total_osadki_night_2, 'Ос.Д' : total_osadki_day_2}
    Day_3 = {'Max' : max_day3,'Min' : min_night3, 'Нv' : max_wind_night_3, 'Дv': max_wind_day_3, 'Ос.Н' : total_osadki_night_3, 'Ос.Д' : total_osadki_day_3}

    InfoForeca = [Day_1, Day_2, Day_3]

    
    return InfoForeca

def create_xlsx(dfs, list_city_name):
    multiple_dfs(dfs, 'foreca', 'InfoWeather.xlsx', 1, list_city_name)
    int_name('InfoWeather.xlsx', 'foreca', list_city_name)


def main():
    list_city_foreca = [
        {'Name' : 'Южно-Сахалинск', 'Url' : 'https://www.foreca.ru/Russia/Yuzhno-Sakhalinsk?details=20231113'},
        {'Name' : 'Холмск', 'Url' : 'https://www.foreca.ru/Russia/Kholmsk?details=20231113'},
        {'Name' : 'Корсаков', 'Url' : 'https://www.foreca.ru/Russia/Korsakov?details=20231113'},
        {'Name' : 'Александровск-Сахалинский', 'Url' : 'https://www.foreca.ru/Russia/Aleksandrovsk_Sakhalinskiy?details=20231113'},
        {'Name' : 'Тымовск', 'Url' : 'https://www.foreca.ru/Russia/Tymovskoye?details=20231113'},
        {'Name' : 'Поронайск', 'Url' : 'https://www.foreca.ru/Russia/Poronaysk?details=20231113'},
        {'Name' : 'Северо-Курильск', 'Url' : "https://www.foreca.ru/Russia/Severo-Kuril'sk?details=20231113"},
        {'Name' : 'Курильск', 'Url' : "https://www.foreca.ru/Russia/Kuril'sk?details=20231113"},
        {'Name' : 'Оха', 'Url' : 'https://www.foreca.ru/Russia/Okha?details=20231113'},
        {'Name' : 'Ноглики', 'Url' : 'https://www.foreca.ru/Russia/Nogliki?details=20231113'},
        {'Name' : 'Южно-Курильск', 'Url' : 'https://www.foreca.ru/Russia/Yuzhno-Kurilsk?details=20231113'},
        {'Name' : 'Углегорск', 'Url' : 'https://www.foreca.ru/Russia/Uglegorsk?details=20231114'},
        {'Name' : 'Ильинск', 'Url' : "https://www.foreca.ru/Russia/Il'inskiy?details=20231114"}
    ]
    list_city_new = []
    list_city_name = []
    dfs = []
    for c in list_city_foreca:
        ci = City(c['Name'], c['Url'])
        list_city_new.append(ci)
        list_city_name.append(str(c['Name']))
    clear = True
    progress = IncrementalBar('Foreca', max = len(list_city_foreca))
    for c in list_city_new:
        progress.next()
        info = get_foreca(c.url)
        dfs.append(table_panda(listinfo = info, name = c.name, check = clear))
        clear = False
    progress.finish()
    return dfs, list_city_name


if __name__ == '__main__':
    main()