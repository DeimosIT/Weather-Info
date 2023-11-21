from selenium import webdriver
from selenium.webdriver.common.by import By
from progress.bar import IncrementalBar
import time
import re
import pandas as pd
from openpyxl import load_workbook
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as EC
from datetime import date
import os

class City:
    def __init__(self, name, url):
        self.name = name
        self.url = url

    def getName(self):
        return self.name
    
    def getUrl(self):
        return self.url


def get_url(url):
    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.page_load_strategy = 'none'
    options.add_experimental_option("prefs", {"profile.managed_default_content_settings.images": 2})
    driver = webdriver.Chrome(options = options)
    driver.get(url)
    
    return driver

def osadki_rain(list):
    rain_total = []
    snow_total = []
    for rain in list:
        string_vr = str(rain)
        num = re.findall(r'[-+]?(?:\d*\.*\d*.мм)', string_vr)
        string_vr2 = str(num)
        num2 = re.findall(r'[-+]?(?:\d*\.*\d+)', string_vr2)
        rain_total.append(float(num2[0]))
    result = round(sum(rain_total), 2)
    
    for snow in list:
        string_vr = str(snow)
        num = re.findall(r'[-+]?(?:\d*\.*\d*.см)', string_vr)
        string_vr2 = str(num)
        num2 = re.findall(r'[-+]?(?:\d*\.*\d+)', string_vr2)
        snow_total.append(float(num2[0]))
    
    result1 = round(sum(snow_total), 2)
    result_total = round(result + result1, 3)
    return result_total   

def table_panda(listinfo):
    df = pd.DataFrame(listinfo, columns=["Min", "Max", "Нv", "Дv", "Ос.Н", "Ос.Д"], index=["Сутки 1", "Сутки 2", "Сутки 3"])
    return df

def multiple_dfs(df_list, sheets, file_name, spaces):
    with pd.ExcelWriter(file_name, mode='a',if_sheet_exists="overlay", engine="openpyxl") as writer:
        row = 0
        for dataframe in zip(df_list):
            row +=1
            dataframe.to_excel(writer,sheet_name=sheets,startrow=row , startcol=0)   
            row = row + len(dataframe.index) + spaces + 1
        writer._save()

def int_name(file_name, sheets, list_name):
    wb = load_workbook(file_name)   
    today = date.today()
    date_3 = today.strftime("%d.%m.%y")
    ws = wb[sheets]
    row = 1
    coll = 1
    for name in list_name:
        ws.cell(row,coll,name)
        ws.cell(row, 7, date_3)
        ws.cell(row, 6, 'Сегодня')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 6

    wb.save(file_name)    

def get_wind(list_wind, list_direct):

    maxWind = 0
    maxDirect = ' '
    for wind, direct in zip(list_wind, list_direct):
        if(wind > maxWind):
            maxWind = wind
            maxDirect = direct
    stringMax = str(maxWind) + str(maxDirect)
    return str.lower(stringMax)

def get_rp5_night(url):
    driver = get_url(url)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'ftab-0'))).click()
    time.sleep(2)
    segodnya = driver.find_element(By.ID, 'forecastTable_1_3')
    hour_day = driver.find_elements(By.XPATH, ('//*[@id="forecastTable_1_3"]/tbody/tr[2]/td[contains(@class, "underlineRow")]'))
    temper_day = segodnya.find_elements(By.CSS_SELECTOR, ('td[dir = ltr]'))
    wind_day = driver.find_elements(By.XPATH, ('//*[@id="forecastTable_1_3"]/tbody/tr[10]/td/div[contains(@class, "wv_0")]'))
    elem = driver.find_elements(By.XPATH, ('//*[@id="forecastTable_1_3"]/tbody/tr[4]/td/div[contains(@class, "pr_0")]'))
    directs = driver.find_elements(By.XPATH, ('//*[@id="forecastTable_1_3"]/tbody/tr[11]/td'))
    
    direct_day_list = []
    for direct in directs:
        direct_day_list.append(direct.text)
    start = 1
    stop = 78
    slice_obj = slice(start, stop)
    direct_day_list_new = direct_day_list[slice_obj]

    hour_day_list = []
    for hour in hour_day:
        hour_day_list.append(hour.text)
    start = 1
    stop = 79
    slice_obj = slice(start, stop)
    hour_day_new = hour_day_list[slice_obj]

    temper_day_list = []
    for temper in temper_day:
        temper_day_list.append(temper.text)    
    stop = 78
    slice_obj = slice(stop)
    temper_day_new = temper_day_list[slice_obj]

    wind_day_list = []
    for wind in wind_day:
        if wind.text == '':
            wind_day_list.append('0')
        else:
            wind_day_list.append(wind.text)    
    stop = 78
    slice_obj = slice(stop)
    wind_day_new = wind_day_list[slice_obj]
    
    osadki_list = []
    for el in elem:
        attrib = el.get_attribute('onmouseover')
        string = attrib
        string += "['0 мм', '0 см']"
        num = re.findall(r'[-+]?(?:\d*\.*\d+).см|[-+]?(?:\d*\.*\d+).мм', string)
        osadki_list.append(num)

    driver.close()
    driver.quit()

    day1_hour = []
    day2_hour = []
    day3_hour = []
    day4_hour = []
    day1_temp = []
    day2_temp = []
    day3_temp = []
    day4_temp = []
    day1_wind = []
    day2_wind = []
    day3_wind = []
    day4_wind = []
    day1_osad = []
    day2_osad = []
    day3_osad = []
    day4_osad = []
    day1_direct = []
    day2_direct = []
    day3_direct = []
    day4_direct = []
    
    night_temper_1 = []
    night_temper_2 = []
    night_temper_3 = []
    night_hour_1 = []
    night_hour_2 = []
    night_hour_3 = []
    night_wind_1 = []
    night_wind_2 = []
    night_wind_3 = []
    night_osad_1 = []
    night_osad_2 = []
    night_osad_3 = []
    night_direct_1 = []
    night_direct_2 = []
    night_direct_3 = []

    day_temper_1 = []
    day_temper_2 = []
    day_temper_3 = []
    day_hour_1 = []
    day_hour_2 = []
    day_hour_3 = []
    day_wind_1 = []
    day_wind_2 = []
    day_wind_3 = []
    day_osad_1 = []
    day_osad_2 = []
    day_osad_3 = []
    day_direct_1 = []
    day_direct_2 = []
    day_direct_3 = []
    

    check = 0
    check_true1 = True
    check_true2 = False
    check_true3 = False
    check_true4 = False

    for temper, hour, wind, osad, direct in zip(temper_day_new, hour_day_new, wind_day_new, osadki_list, direct_day_list_new):
        num_hour = int(hour)
        if(check < num_hour):
            
            if(check_true1):
                day1_hour.append(hour)
                day1_temp.append(temper)
                day1_wind.append(wind)
                day1_osad.append(osad)
                day1_direct.append(direct)
            if(check_true2):
                day2_hour.append(hour)
                day2_temp.append(temper)
                day2_wind.append(wind)
                day2_osad.append(osad)
                day2_direct.append(direct)
            if(check_true3):
                day3_hour.append(hour)
                day3_temp.append(temper)
                day3_wind.append(wind)
                day3_osad.append(osad)
                day3_direct.append(direct)
            if(check_true4):
                day4_hour.append(hour)
                day4_temp.append(temper)
                day4_wind.append(wind)
                day4_osad.append(osad)
                day4_direct.append(direct)
        else:
            if(check_true2 == False and check_true3 == False and check_true4 == False):
                check_true1 = False
                check_true2 = True
                day2_hour.append(hour)
                day2_temp.append(temper)
                day2_wind.append(wind)
                day2_osad.append(osad)
                day2_direct.append(direct)
                check = num_hour  
                continue
            if(check_true3 == False and check_true1 == False and check_true4 == False):     
                check_true3 = True
                check_true2 = False
                day3_hour.append(hour)
                day3_temp.append(temper)
                day3_wind.append(wind)
                day3_osad.append(osad)
                day3_direct.append(direct)
                check = num_hour
                continue
            if(check_true4 == False and check_true1 == False and check_true2 == False):
                check_true4 = True
                check_true3 = False
                day4_hour.append(hour)
                day4_temp.append(temper)
                day4_wind.append(wind)
                day4_osad.append(osad)
                day4_direct.append(direct)
                check = num_hour
                continue
            if(check_true4 == True):
                break
        check = num_hour
    
    for temper, hour, wind, osad, direct in zip(day1_temp, day1_hour, day1_wind, day1_osad, day1_direct):
        if(int(hour) >= 20):
            night_hour_1.append(int(hour))
            night_temper_1.append(int(temper))
            night_wind_1.append(int(wind))
            night_direct_1.append(direct)
        if(int(hour) >= 21):
            night_osad_1.append(osad)
    for temper, hour, wind, osad, direct in zip(day2_temp, day2_hour, day2_wind, day2_osad, day2_direct):
        if(int(hour) <= 8):
            night_hour_1.append(int(hour))
            night_temper_1.append(int(temper))
            night_wind_1.append(int(wind))
            night_osad_1.append(osad)
            night_direct_1.append(direct)
        if(int(hour) >= 20):
            night_hour_2.append(int(hour))
            night_temper_2.append(int(temper))
            night_wind_2.append(int(wind))
            night_direct_2.append(direct)
        if(int(hour) >= 21):
            night_osad_2.append(osad)
        if(int(hour) >=9 and int(hour) <= 19):
            day_hour_1.append(int(hour))
            day_temper_1.append(int(temper))
            day_wind_1.append(int(wind))
            day_direct_1.append(direct)
        if(int(hour) >=9 and int(hour) <= 20):
            day_osad_1.append(osad)
    for temper, hour, wind, osad, direct in zip(day3_temp, day3_hour, day3_wind, day3_osad, day3_direct):
        if(int(hour) <= 8):
            night_hour_2.append(int(hour))
            night_temper_2.append(int(temper))
            night_wind_2.append(int(wind))
            night_osad_2.append(osad)
            night_direct_2.append(direct)
            if (int(hour) == 8):
                x = 0
                while x!=2:
                    night_osad_2.append(osad)
                    x += 1
        if(int(hour) >= 20):
            night_hour_3.append(int(hour))
            night_temper_3.append(int(temper))
            night_wind_3.append(int(wind))
            night_direct_3.append(direct)
        if(int(hour) >= 21):
            night_osad_3.append(osad)
        if(int(hour) >=9 and int(hour) <= 19):
            day_hour_2.append(int(hour))
            day_temper_2.append(int(temper))
            day_wind_2.append(int(wind))
            day_direct_2.append(direct)
        if(int(hour) >=9 and int(hour) <= 20):
            day_osad_2.append(osad)
            if(int(hour) >=11):
                x = 0
                while x!=2:
                    day_osad_2.append(osad)
                    x += 1
    for temper, hour, wind, osad, direct in zip(day4_temp, day4_hour, day4_wind, day4_osad, day4_direct):
        if(int(hour) <= 8):
            night_hour_3.append(int(hour))
            night_temper_3.append(int(temper))
            night_wind_3.append(int(wind))
            night_direct_3.append(direct)
            x = 0
            while x != 3:
                night_osad_3.append(osad)
                x += 1
        if(int(hour) >=9 and int(hour) <= 19):
            day_hour_3.append(int(hour))
            day_temper_3.append(int(temper))
            day_wind_3.append(int(wind))
            day_direct_3.append(direct)
        if(int(hour) >= 10 and int(hour) <=20):
            x = 0
            while x != 3:
                day_osad_3.append(osad)
                x += 1

    total_osadki_night_1 = osadki_rain(night_osad_1)
    total_osadki_night_2 = osadki_rain(night_osad_2)
    total_osadki_night_3 = osadki_rain(night_osad_3)
    total_osadki_day_1 = osadki_rain(day_osad_1)
    total_osadki_day_2 = osadki_rain(day_osad_2)
    total_osadki_day_3 = osadki_rain(day_osad_3)

    min_night1 = min(night_temper_1)
    min_night2 = min(night_temper_2)
    min_night3 = min(night_temper_3)
    max_day1 = max(day_temper_1)
    max_day2 = max(day_temper_2)
    max_day3 = max(day_temper_3)
    max_wind_night_1 = get_wind(night_wind_1, night_direct_1)
    max_wind_night_2 = get_wind(night_wind_2, night_direct_2)
    max_wind_night_3 = get_wind(night_wind_3, night_direct_3)
    max_wind_day_1 = get_wind(day_wind_1, day_direct_1)
    max_wind_day_2 = get_wind(day_wind_2, day_direct_2)
    max_wind_day_3 = get_wind(day_wind_3, day_direct_3)

    Day_1 = {
        'Max' : max_day1,
        'Min' : min_night1,
        'Нv' : max_wind_night_1,
        'Дv': max_wind_day_1,
        'Ос.Н' : total_osadki_night_1,
        'Ос.Д' : total_osadki_day_1
        }
    Day_2 = {'Max' : max_day2,'Min' : min_night2, 'Нv' : max_wind_night_2, 'Дv': max_wind_day_2, 'Ос.Н' : total_osadki_night_2, 'Ос.Д' : total_osadki_day_2}
    Day_3 = {'Max' : max_day3,'Min' : min_night3, 'Нv' : max_wind_night_3, 'Дv': max_wind_day_3, 'Ос.Н' : total_osadki_night_3, 'Ос.Д' : total_osadki_day_3}

    InfoRP5 = [Day_1, Day_2, Day_3]

    
    return InfoRP5

def create_xlsx(dfs, list_city_name):
    multiple_dfs(dfs, 'RP5', 'InfoWeather.xlsx', 1, list_city_name)
    int_name('InfoWeather.xlsx', 'RP5', list_city_name)

def main():

    list_city_rp5 = [
        {'Name' : 'Южно-Сахалинск', 'Url' : 'https://rp5.ru/Погода_в_Южно-Сахалинске'},
        {'Name' : 'Холмск', 'Url' : 'https://rp5.ru/Погода_в_Холмске'},
        {'Name' : 'Корсаков', 'Url' : 'https://rp5.ru/Погода_в_Корсакове'},
        {'Name' : 'Александровск-Сахалинский', 'Url' : 'https://rp5.ru/Погода_в_Александровске-Сахалинском'},
        {'Name' : 'Тымовск', 'Url' : 'https://rp5.ru/Погода_в_Тымовском'},
        {'Name' : 'Поронайск', 'Url' : 'https://rp5.ru/Погода_в_Поронайске'},
        {'Name' : 'Северо-Курильск', 'Url' : 'https://rp5.ru/Погода_в_Северо-Курильске'},
        {'Name' : 'Курильск', 'Url' : 'https://rp5.ru/Погода_в_Курильске'},
        {'Name' : 'Оха', 'Url' : 'https://rp5.ru/Погода_в_Охе'},
        {'Name' : 'Ноглики', 'Url' : 'https://rp5.ru/Погода_в_Ногликах'},
        {'Name' : 'Южно-Курильск', 'Url' : 'https://rp5.ru/Погода_в_Южно-Курильске'},
        {'Name' : 'Углегорск', 'Url' : 'https://rp5.ru/Погода_в_Углегорске,_Сахалинская_область'},
        {'Name' : 'Ильинск', 'Url' : 'https://rp5.ru/Погода_в_Ильинском,_Сахалинская_область'}
    ]
    list_city_new = []
    list_city_name = []
    dfs = []
    for c in list_city_rp5:
        ci = City(c['Name'], c['Url'])
        list_city_new.append(ci)
        list_city_name.append(str(c['Name']))
    clear = True
    progress = IncrementalBar('RP5', max = len(list_city_rp5))
    for c in list_city_new:
        progress.next()
        info = get_rp5_night(c.url)
        dfs.append(table_panda(listinfo = info))
        clear = False
    progress.finish()
    return dfs, list_city_name

    

if __name__ == '__main__':
    main()
    